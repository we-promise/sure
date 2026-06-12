# Generates a synthetic Crédit Mutuel / CIC-style multi-account .xlsx export
# used as a test fixture for XlsxImport. No personal data — fake account numbers.
#
# Run from repo root:  python test/fixtures/files/imports/generate_bank_export.py
import datetime
from openpyxl import Workbook

wb = Workbook()

# 1) Summary sheet ----------------------------------------------------------
ws = wb.active
ws.title = "Vos comptes"
ws["A1"] = "Votre situation financière au 12/06/2026"
ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "Compte", "R.I.B.", "Solde", "Dev"
ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "Compte Courant Test", "10278 02619 00099999901", 100.0, "EUR"
ws["A5"], ws["B5"], ws["C5"], ws["D5"] = "Livret Test", "10278 02619 00099999902", 500.0, "EUR"

# 2) Current account ("cpt") ------------------------------------------------
cpt = wb.create_sheet("Cpt 02619 00099999901")
cpt["A1"] = "Situation de votre compte Compte Courant Test (EUR) au 12/06/2026"
cpt["A2"] = "R.I.B. : 10278 02619 00099999901"
cpt["A4"] = "Liste de vos comptes"
cpt["A5"], cpt["B5"], cpt["C5"] = "Date", "Valeur", "Libellé"
cpt["D5"], cpt["E5"], cpt["F5"], cpt["G5"] = "Débit", "Crédit", "Solde", "Dev"
# Date, Valeur, Libellé, Débit, Crédit, Solde, Dev
cpt_rows = [
    (datetime.date(2025, 12, 31), datetime.date(2026, 1, 1), "INTERETS 2025", None, 12.34, None, "EUR"),
    (datetime.date(2026, 1, 15), datetime.date(2026, 1, 15), "PAIEMENT CB CARREFOUR", -45.67, None, None, "EUR"),
    (datetime.date(2026, 2, 1), datetime.date(2026, 2, 1), "NOUVEAU TAUX", None, 0, None, "EUR"),  # zero -> skipped
]
for i, row in enumerate(cpt_rows, start=6):
    for j, val in enumerate(row):
        cpt.cell(row=i, column=j + 1, value=val)

# 3) Card account ("cb") — same account number as the cpt above --------------
cb = wb.create_sheet("CB 02619 00099999901 # 052026")
cb["A1"] = "Encours prélevé : Fin mai 2026 : - 74,06 EUR"
cb["A2"] = "R.I.B. : 10278 02619 00099999901 - 05/2026"
cb["A4"] = "Liste de vos comptes"
cb["A5"], cb["B5"], cb["C5"], cb["D5"] = "Date", "Libellé", "Montant", "Dev"
cb["A6"] = "Carte Mastercard XXXXXXXXXXXX9999 (EUR)"  # sub-header row, skipped via range
cb_rows = [
    (datetime.date(2026, 5, 18), "OPENAI CHATGPT", -23.0, "EUR"),
    (datetime.date(2026, 5, 17), "LIDL", -38.56, "EUR"),
    (datetime.date(2026, 5, 16), "JARDILAND", -12.5, "EUR"),
]
for i, row in enumerate(cb_rows, start=7):
    for j, val in enumerate(row):
        cb.cell(row=i, column=j + 1, value=val)

# 4) Hidden manifest --------------------------------------------------------
manifest = wb.create_sheet("hidden_data")
manifest.sheet_state = "veryHidden"
manifest["A1"] = "Nombre de compte"
manifest["A2"] = 2
manifest["A3"], manifest["B3"], manifest["C3"] = "Nom du compte", "Range", "Type"
manifest["A4"], manifest["B4"], manifest["C4"] = "Cpt 02619 00099999901", "A6:E8", "cpt"
manifest["A5"], manifest["B5"], manifest["C5"] = "CB 02619 00099999901 # 052026", "A7:C9", "cb"

out = "test/fixtures/files/imports/sample_bank_export.xlsx"
wb.save(out)
print("wrote", out)
